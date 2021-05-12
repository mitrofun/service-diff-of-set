from typing import Union

from loguru import logger
import xlrd  # type: ignore
import openpyxl  # type: ignore


class ExcelReader:
    before_list: list = []
    after_list: list = []

    def __init__(self, filename):
        self.filename = filename

    def lists_for_comparison(self):
        raise NotImplementedError


class XlsxReader(ExcelReader):

    @staticmethod
    def _get_target_list(col_title: str, sheet, headers: tuple) -> list:
        result = []
        col_number = headers.index(col_title) + 1
        for row_number in range(2, sheet.max_row + 1):
            result.append(sheet.cell(row_number, col_number).value)
        return result

    @staticmethod
    def _get_headers(sheet):
        return next(sheet.values)

    def lists_for_comparison(self) -> tuple[list[Union[int, None]], list[Union[int, None]]]:
        wb = openpyxl.load_workbook(filename=self.filename)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            try:
                headers = self._get_headers(ws)
                if 'before' and 'after' not in headers:
                    return self.before_list, self.after_list
                self.before_list = self._get_target_list('before', ws, headers)
                self.after_list = self._get_target_list('after', ws, headers)
            except StopIteration:
                logger.warning(f'Clear worksheet "{sheet_name}" in file: {self.filename}.')
        wb.close()
        return self.before_list, self.after_list


class XlsReader(ExcelReader):

    @staticmethod
    def _get_target_list(col_title: str, sheet, headers: list) -> list:
        idx = headers.index(col_title)
        return [int(i) for i in sheet.col_values(idx)[1:] if i != '']

    @staticmethod
    def _get_headers(sheet):
        return list(map(lambda x: x.value, sheet.row(0)))

    def lists_for_comparison(self) -> tuple[list[Union[int, None]], list[Union[int, None]]]:
        wb = xlrd.open_workbook(self.filename)
        for sheet in wb.sheets():
            ws = wb[sheet.name]
            try:
                headers = self._get_headers(ws)
                if 'before' and 'after' not in headers:
                    return self.before_list, self.after_list
                self.before_list = self._get_target_list('before', ws, headers)
                self.after_list = self._get_target_list('after', ws, headers)
            except IndexError:
                logger.warning(f'Clear worksheet "{sheet.name}" in file: {self.filename}.')
        return self.before_list, self.after_list


READERS = {
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': XlsxReader,
    'application/vnd.ms-excel': XlsReader,
}


def get_lists_for_comparison(content_type: str,
                             filename: str) -> tuple[list[Union[int, None]], list[Union[int, None]]]:
    try:
        reader = READERS[content_type](filename)
        return reader.lists_for_comparison()
    except KeyError:
        logger.warning(f'Not supported content type "{content_type}" in file: {filename}.')
        return [], []
